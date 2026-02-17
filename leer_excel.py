"""
Lector de Excel - KPI Perforación y Tronadura (PyT)

Estructura del archivo Excel:
- Cada hoja es una fecha (ej: '15-01', '31-01')
- Hojas especiales: 'Datos ', 'Patios' (se excluyen)

Estructura de cada hoja (METROS DE PERFORACION):
  Fila 4:  Encabezados (Turno A, Turno B, Total, Plan, Cumplimientos, Estados)
  Filas 6-9:   Sección F12 (Fase 12) - TOTAL en fila 9
  Filas 11-15:  Sección F10 (Fase 10) - TOTAL en fila 15
  Filas 17-25:  Sección F09 (Fase 9)  - TOTAL en fila 25
  Filas 27-31:  Sección F11 (Fase 11) - TOTAL en fila 31
  Fila 34:  TOTAL METROS (todas las fases)
  Filas 38-41:  METROS ROC por fase (F09, F10, F11, F12)
  Fila 42:  PLAN DIARIO ROC

Columnas principales:
  B: Fase (F12, F10, F09, F11)
  C: Equipo (PF03, PF07, PF21, PF22, PF23, PF24, PF25, PF26, ROC, PFARR)
  D: Turno A
  E: Turno B
  F: Total
  G: Plan
  H: Cumplimiento Turno A
  I: Cumplimiento Turno B
  K: Cumplimiento Diario
  M: Estado Perforadora Turno A
  N: Estado Perforadora Turno B
"""

import openpyxl
from dataclasses import dataclass, field
from typing import Optional
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# Ruta por defecto del archivo Excel
EXCEL_PATH = "KPI PyT  Enero 2026 NO OFICIAL.xlsx"

# Hojas que NO son fechas (se excluyen del procesamiento)
HOJAS_EXCLUIDAS = {"Datos", "Patios"}

# Definición de secciones de fases con sus rangos de filas
# (fase_label, fila_inicio_equipos, fila_total)
SECCIONES_FASES = [
    ("F12", 6, 9),
    ("F10", 11, 15),
    ("F09", 17, 25),
    ("F11", 27, 31),
]

# Fila de totales generales
FILA_TOTAL_METROS = 34

# Metros ROC
FILA_METROS_ROC_HEADER = 37
FILAS_METROS_ROC = {
    "Fase 9": 38,
    "Fase 10": 39,
    "Fase 11": 40,
    "Fase 12": 41,
}
FILA_PLAN_DIARIO_ROC = 42


@dataclass
class EquipoData:
    """Datos de un equipo (perforadora) en una fase."""
    equipo: str
    turno_a: float = 0.0
    turno_b: float = 0.0
    total: float = 0.0
    plan: float = 0.0
    cumplimiento_ta: float = 0.0
    cumplimiento_tb: float = 0.0
    cumplimiento_diario: float = 0.0
    estado_ta: str = ""
    estado_tb: str = ""


@dataclass
class FaseData:
    """Datos de una fase completa con sus equipos y totales."""
    fase: str
    equipos: list = field(default_factory=list)
    total_turno_a: float = 0.0
    total_turno_b: float = 0.0
    total: float = 0.0
    plan: float = 0.0
    cumplimiento_ta: float = 0.0
    cumplimiento_tb: float = 0.0
    cumplimiento_diario: float = 0.0


@dataclass
class MetrosRocData:
    """Datos de metros ROC por fase."""
    fase_9: float = 0.0
    fase_10: float = 0.0
    fase_11: float = 0.0
    fase_12: float = 0.0
    plan_diario: float = 0.0
    turno_a: float = 0.0
    turno_b: float = 0.0
    total: float = 0.0


@dataclass
class DiaData:
    """Datos completos de un día (hoja del Excel)."""
    fecha: str
    fases: dict = field(default_factory=dict)  # {nombre_fase: FaseData}
    total_turno_a: float = 0.0
    total_turno_b: float = 0.0
    total_metros: float = 0.0
    plan_total: float = 0.0
    cumplimiento_ta: float = 0.0
    cumplimiento_tb: float = 0.0
    cumplimiento_diario: float = 0.0
    metros_roc: MetrosRocData = field(default_factory=MetrosRocData)


def _safe_float(value) -> float:
    """Convierte un valor a float de forma segura."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(value) -> str:
    """Convierte un valor a string de forma segura."""
    if value is None:
        return ""
    return str(value).strip()


def _leer_equipo(sheet, fila: int) -> Optional[EquipoData]:
    """Lee los datos de un equipo en una fila específica."""
    nombre = _safe_str(sheet[f"C{fila}"].value)
    if not nombre or nombre == "TOTAL":
        return None

    return EquipoData(
        equipo=nombre,
        turno_a=_safe_float(sheet[f"D{fila}"].value),
        turno_b=_safe_float(sheet[f"E{fila}"].value),
        total=_safe_float(sheet[f"F{fila}"].value),
        plan=_safe_float(sheet[f"G{fila}"].value),
        cumplimiento_ta=_safe_float(sheet[f"H{fila}"].value),
        cumplimiento_tb=_safe_float(sheet[f"I{fila}"].value),
        cumplimiento_diario=_safe_float(sheet[f"K{fila}"].value),
        estado_ta=_safe_str(sheet[f"M{fila}"].value),
        estado_tb=_safe_str(sheet[f"N{fila}"].value),
    )


def _leer_fase(sheet, fase_label: str, fila_inicio: int, fila_total: int) -> FaseData:
    """Lee todos los datos de una sección de fase."""
    fase = FaseData(fase=fase_label)

    # Leer equipos (desde fila_inicio hasta fila_total - 1)
    for fila in range(fila_inicio, fila_total):
        equipo = _leer_equipo(sheet, fila)
        if equipo:
            fase.equipos.append(equipo)

    # Leer totales de la fase
    fase.total_turno_a = _safe_float(sheet[f"D{fila_total}"].value)
    fase.total_turno_b = _safe_float(sheet[f"E{fila_total}"].value)
    fase.total = _safe_float(sheet[f"F{fila_total}"].value)
    fase.plan = _safe_float(sheet[f"G{fila_total}"].value)
    fase.cumplimiento_ta = _safe_float(sheet[f"H{fila_total}"].value)
    fase.cumplimiento_tb = _safe_float(sheet[f"I{fila_total}"].value)
    fase.cumplimiento_diario = _safe_float(sheet[f"K{fila_total}"].value)

    return fase


def _leer_metros_roc(sheet) -> MetrosRocData:
    """Lee la sección de Metros ROC."""
    roc = MetrosRocData()

    # Fase 9 - fila 38
    roc.fase_9 = _safe_float(sheet[f"F{FILAS_METROS_ROC['Fase 9']}"].value)
    # Fase 10 - fila 39
    roc.fase_10 = _safe_float(sheet[f"F{FILAS_METROS_ROC['Fase 10']}"].value)
    # Fase 11 - fila 40
    roc.fase_11 = _safe_float(sheet[f"F{FILAS_METROS_ROC['Fase 11']}"].value)
    # Fase 12 - fila 41
    roc.fase_12 = _safe_float(sheet[f"F{FILAS_METROS_ROC['Fase 12']}"].value)

    # Plan diario ROC - fila 42
    roc.plan_diario = _safe_float(sheet[f"C{FILA_PLAN_DIARIO_ROC}"].value)
    roc.turno_a = _safe_float(sheet[f"D{FILA_PLAN_DIARIO_ROC}"].value)
    roc.turno_b = _safe_float(sheet[f"E{FILA_PLAN_DIARIO_ROC}"].value)
    roc.total = _safe_float(sheet[f"F{FILA_PLAN_DIARIO_ROC}"].value)

    return roc


def leer_dia(wb, nombre_hoja: str) -> DiaData:
    """Lee todos los datos de una hoja (un día)."""
    sheet = wb[nombre_hoja]
    dia = DiaData(fecha=nombre_hoja.strip())

    # Leer cada sección de fase
    for fase_label, fila_inicio, fila_total in SECCIONES_FASES:
        fase = _leer_fase(sheet, fase_label, fila_inicio, fila_total)
        dia.fases[fase_label] = fase

    # Leer TOTAL METROS (fila 34)
    dia.total_turno_a = _safe_float(sheet[f"D{FILA_TOTAL_METROS}"].value)
    dia.total_turno_b = _safe_float(sheet[f"E{FILA_TOTAL_METROS}"].value)
    dia.total_metros = _safe_float(sheet[f"F{FILA_TOTAL_METROS}"].value)
    dia.plan_total = _safe_float(sheet[f"G{FILA_TOTAL_METROS}"].value)
    dia.cumplimiento_ta = _safe_float(sheet[f"H{FILA_TOTAL_METROS}"].value)
    dia.cumplimiento_tb = _safe_float(sheet[f"I{FILA_TOTAL_METROS}"].value)
    dia.cumplimiento_diario = _safe_float(sheet[f"K{FILA_TOTAL_METROS}"].value)

    # Leer Metros ROC
    dia.metros_roc = _leer_metros_roc(sheet)

    return dia


def cargar_excel(ruta: str = EXCEL_PATH) -> dict:
    """
    Carga el archivo Excel completo y retorna un diccionario con todos los días.

    Returns:
        dict con claves:
            - 'dias': {nombre_hoja: DiaData} para cada hoja de fecha
            - 'hojas_fecha': lista de nombres de hojas que son fechas
            - 'total_hojas': cantidad de hojas procesadas
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)

    hojas_fecha = [name for name in wb.sheetnames if name.strip() not in HOJAS_EXCLUIDAS]
    dias = {}

    for nombre in hojas_fecha:
        dias[nombre.strip()] = leer_dia(wb, nombre)

    wb.close()

    return {
        "dias": dias,
        "hojas_fecha": [h.strip() for h in hojas_fecha],
        "total_hojas": len(hojas_fecha),
    }


def obtener_dia(fecha: str, ruta: str = EXCEL_PATH) -> Optional[DiaData]:
    """
    Carga y retorna los datos de un solo día.

    Args:
        fecha: nombre de la hoja (ej: '15-01', '31-12')
        ruta: ruta al archivo Excel
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)

    # Buscar la hoja (puede tener espacios al final)
    nombre_hoja = None
    for name in wb.sheetnames:
        if name.strip() == fecha.strip():
            nombre_hoja = name
            break

    if nombre_hoja is None:
        wb.close()
        return None

    dia = leer_dia(wb, nombre_hoja)
    wb.close()
    return dia


def imprimir_dia_plano(dia: DiaData):
    """Imprime los datos en formato tabla plana: 1 equipo por fila."""
    print(f"\n{'='*90}")
    print(f"  METROS DE PERFORACION (Tabla Plana) - {dia.fecha}")
    print(f"{'='*90}")
    print(
        f"{'Fecha':<8} {'Fase':<6} {'Equipo':<8} "
        f"{'Turno A':>9} {'Turno B':>9} {'Total':>9} "
        f"{'Plan':>9} {'Cump.Dia':>10}"
    )
    print(f"{'-'*90}")

    for fase_label in ["F12", "F10", "F09", "F11"]:
        fase = dia.fases.get(fase_label)
        if not fase:
            continue
        for eq in fase.equipos:
            print(
                f"{dia.fecha:<8} {fase_label:<6} {eq.equipo:<8} "
                f"{eq.turno_a:>9.1f} {eq.turno_b:>9.1f} {eq.total:>9.1f} "
                f"{eq.plan:>9.1f} {eq.cumplimiento_diario:>9.1%}"
            )

    print(f"{'-'*90}")
    print(
        f"{dia.fecha:<8} {'TOTAL':<6} {'':8} "
        f"{dia.total_turno_a:>9.1f} {dia.total_turno_b:>9.1f} {dia.total_metros:>9.1f} "
        f"{dia.plan_total:>9.1f} {dia.cumplimiento_diario:>9.1%}"
    )


def imprimir_dia(dia: DiaData):
    """Imprime un resumen formateado de un día."""
    print(f"\n{'='*70}")
    print(f"  METROS DE PERFORACION - {dia.fecha}")
    print(f"{'='*70}")
    print(f"{'Fase':<6} {'Equipo':<8} {'Turno A':>9} {'Turno B':>9} {'Total':>9} {'Plan':>9} {'Cump.TA':>9} {'Cump.TB':>9} {'Cump.Dia':>9}")
    print(f"{'-'*70}")

    for fase_label in ["F12", "F10", "F09", "F11"]:
        fase = dia.fases.get(fase_label)
        if not fase:
            continue

        for i, eq in enumerate(fase.equipos):
            fase_col = fase_label if i == 0 else ""
            print(
                f"{fase_col:<6} {eq.equipo:<8} "
                f"{eq.turno_a:>9.1f} {eq.turno_b:>9.1f} {eq.total:>9.1f} "
                f"{eq.plan:>9.1f} {eq.cumplimiento_ta:>8.1%} {eq.cumplimiento_tb:>8.1%} "
                f"{eq.cumplimiento_diario:>8.1%}"
            )
            if eq.estado_ta or eq.estado_tb:
                print(f"{'':>15}Estado TA: {eq.estado_ta}  |  Estado TB: {eq.estado_tb}")

        if not fase.equipos:
            print(f"{fase_label:<6} {'(sin equipos)':<8}")

        print(
            f"{'':>6}{'TOTAL':<8} "
            f"{fase.total_turno_a:>9.1f} {fase.total_turno_b:>9.1f} {fase.total:>9.1f} "
            f"{fase.plan:>9.1f} {fase.cumplimiento_ta:>8.1%} {fase.cumplimiento_tb:>8.1%} "
            f"{fase.cumplimiento_diario:>8.1%}"
        )
        print()

    print(f"{'-'*70}")
    print(
        f"{'TOTAL METROS':<15}"
        f"{dia.total_turno_a:>9.1f} {dia.total_turno_b:>9.1f} {dia.total_metros:>9.1f} "
        f"{dia.plan_total:>9.1f} {dia.cumplimiento_ta:>8.1%} {dia.cumplimiento_tb:>8.1%} "
        f"{dia.cumplimiento_diario:>8.1%}"
    )

    print(f"\n  METROS ROC:")
    print(f"    Fase 9:  {dia.metros_roc.fase_9:>9.1f}")
    print(f"    Fase 10: {dia.metros_roc.fase_10:>9.1f}")
    print(f"    Fase 11: {dia.metros_roc.fase_11:>9.1f}")
    print(f"    Fase 12: {dia.metros_roc.fase_12:>9.1f}")
    print(f"    Plan Diario ROC: {dia.metros_roc.plan_diario:>9.1f}")
    print(f"    ROC Total (TA:{dia.metros_roc.turno_a:.1f} TB:{dia.metros_roc.turno_b:.1f}): {dia.metros_roc.total:.1f}")


# --- Ejecución directa ---
if __name__ == "__main__":
    import sys

    args = sys.argv[1:]
    modo_plano = "--plano" in args
    fechas = [a for a in args if not a.startswith("--")]

    if fechas:
        for fecha in fechas:
            dia = obtener_dia(fecha)
            if dia:
                if modo_plano:
                    imprimir_dia_plano(dia)
                else:
                    imprimir_dia(dia)
            else:
                print(f"No se encontró la hoja '{fecha}' en el Excel.")
                print("Hojas disponibles:")
                wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
                for name in wb.sheetnames:
                    if name.strip() not in HOJAS_EXCLUIDAS:
                        print(f"  - {name.strip()}")
                wb.close()
    else:
        print("Cargando todas las hojas...")
        data = cargar_excel()
        print(f"Total hojas procesadas: {data['total_hojas']}")
        print(f"Fechas disponibles: {', '.join(data['hojas_fecha'][:10])}... (y más)")

        # Mostrar los últimos 3 días como ejemplo
        ultimas = data["hojas_fecha"][-3:]
        for fecha in ultimas:
            imprimir_dia(data["dias"][fecha])
